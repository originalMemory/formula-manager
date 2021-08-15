#!/user/bin/env python3
# coding=utf-8
"""
@project : formula-manager
@ide     : PyCharm
@file    : export_view
@author  : illusion
@desc    : 
@create  : 2021/8/14
"""
import os

from PyQt5.QtGui import QStandardItemModel, QStandardItem, QDoubleValidator
from PyQt5.QtWidgets import QWidget, QMessageBox, QFileDialog
from openpyxl import Workbook

from db_helper import db_helper
from model.formula import Formula
from ui.export import Ui_Export


class ExportView(QWidget, Ui_Export):
    def __init__(self, parent=None):
        super(ExportView, self).__init__(parent)
        self.setupUi(self)

        self.lineEditWeight.setValidator(QDoubleValidator())
        self.tableView.horizontalHeader().setVisible(False)

        self.pushButtonSearch.clicked.connect(self._search)
        self.pushButtonCompute.clicked.connect(self._compute)
        self.pushButtonExport.clicked.connect(self.export_excel)
        self.lineEditCustomName.returnPressed.connect(self._search)
        self.lineEditColorNo.returnPressed.connect(self._search)
        self.lineEditWeight.returnPressed.connect(self._compute)

        self.model = QStandardItemModel()
        self.tableView.setModel(self.model)
        self.formulas = []

    def _search(self):
        custom_name = self.lineEditCustomName.text()
        color_no = self.lineEditColorNo.text()
        if not color_no and not custom_name:
            QMessageBox.information(self, "提示", "客户名称或色号至少要输入1个！", QMessageBox.Ok)
            return
        sql = 'SELECT * FROM formula WHERE 1=1'
        if custom_name:
            res = db_helper.execute_one(f"SELECT formula_ids FROM custom WHERE name like '%{custom_name}%'")
            if res:
                sql += f" AND id IN ({res['formula_ids']})"
            else:
                QMessageBox.information(self, "提示", "找不到对应的客户！", QMessageBox.Ok)
                return
        if color_no:
            sql += f" AND color_no LIKE '%{color_no}%'"
        res = db_helper.execute_all(sql)
        if res:
            self.formulas = [Formula.from_dict(x) for x in res]
            self._load_data()

    def _compute(self):
        if not self.lineEditWeight.text():
            QMessageBox.information(self, "提示", "重量必须输入！", QMessageBox.Ok)
            return
        self._load_data()

    def _load_data(self):
        self.model.clear()
        custom_name = self.lineEditCustomName.text()
        weight_text = self.lineEditWeight.text()
        if weight_text:
            weight = float(weight_text)
        else:
            weight = 0.0
        formula_length = len(self.formulas)
        for i in range(formula_length):
            formula = self.formulas[i]
            self.model.appendRow(
                [QStandardItem('客户'), QStandardItem(custom_name), QStandardItem('色号'),
                 QStandardItem(formula.color_no), QStandardItem('颜色'), QStandardItem(formula.color_name)])
            self.model.appendRow([QStandardItem('品质'), QStandardItem(formula.quality), QStandardItem('重量'),
                                  QStandardItem(weight_text)])

            dye_length = len(formula.dyes)
            for j in range(dye_length):
                dye = formula.dyes[j]
                if j == 0:
                    pre = '染料'
                else:
                    pre = ''
                self.model.appendRow(
                    [QStandardItem(pre), QStandardItem(dye.name), QStandardItem(str(dye.value * weight * 4.54))])
            catalyzer_length = len(formula.catalyzers)
            for j in range(catalyzer_length):
                catalyzer = formula.catalyzers[j]
                if j == 0:
                    pre = '催化剂'
                else:
                    pre = ''
                self.model.appendRow(
                    [QStandardItem(pre), QStandardItem(catalyzer.name), QStandardItem(f'{catalyzer.value}%')])

            if i <= formula_length - 1:
                self.model.appendRow([])

    def export_excel(self):
        file_paths = QFileDialog.getSaveFileName(self, '导出文件', filter="xlsx files (*.xlsx)")
        if not len(file_paths) or not file_paths[0]:
            return
        export_file = file_paths[0]
        workbook = Workbook()

        custom_name = self.lineEditCustomName.text()
        weight_text = self.lineEditWeight.text()
        if weight_text:
            weight = float(weight_text)
        else:
            weight = 0.0
        formula_length = len(self.formulas)
        for i in range(formula_length):
            formula = self.formulas[i]
            if i == 0:
                # workbook.create_sheet()
                sheet = workbook.worksheets[0]
                sheet.title = formula.color_no
            else:
                sheet = workbook.create_sheet(title=formula.color_no)

            sheet.append(['客户', custom_name, '色号', formula.color_no, '颜色', formula.color_name])
            sheet.append(['品质', formula.quality, '重量', str(weight)])

            dye_length = len(formula.dyes)
            for j in range(dye_length):
                dye = formula.dyes[j]
                if j == 0:
                    pre = '染料'
                else:
                    pre = ''
                sheet.append([pre, dye.name, str(dye.value * weight * 4.54)])
            catalyzer_length = len(formula.catalyzers)
            for j in range(catalyzer_length):
                catalyzer = formula.catalyzers[j]
                if j == 0:
                    pre = '催化剂'
                else:
                    pre = ''
                sheet.append([pre, catalyzer.name, f'{catalyzer.value}%'])

        try:
            if os.path.exists(export_file):
                os.remove(export_file)
            workbook.save(export_file)
            res = QMessageBox.information(self, "提示", "保存成功，是否打开保存文件夹？", QMessageBox.Ok | QMessageBox.Cancel)
            if res == QMessageBox.Ok:
                path = os.path.dirname(export_file)
                path = path.replace('/', '\\')
                os.system("explorer %s" % path)

        except Exception as e:
            QMessageBox.information(self, "提示", str(e), QMessageBox.Ok)


